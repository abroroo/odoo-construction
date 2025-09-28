import logging
import base64
import io
from datetime import datetime

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

try:
    import xlrd
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    xlrd = None
    openpyxl = None

_logger = logging.getLogger(__name__)


class SmetaImportLog(models.Model):
    """Log of smeta import operations"""
    _name = 'construction.smeta.import.log'
    _description = 'Smeta Import Log'
    _order = 'create_date desc'
    _rec_name = 'display_name'

    display_name = fields.Char(compute='_compute_display_name', store=True)
    project_id = fields.Many2one('project.project', string='Project', required=True)
    budget_id = fields.Many2one('construction.project.budget', string='Budget')
    filename = fields.Char(string='File Name', required=True)
    import_date = fields.Datetime(string='Import Date', default=fields.Datetime.now)
    user_id = fields.Many2one('res.users', string='Imported By', default=lambda self: self.env.user)
    state = fields.Selection([
        ('draft', 'Draft'),
        ('processing', 'Processing'),
        ('success', 'Success'),
        ('failed', 'Failed'),
    ], string='Status', default='draft')

    # Statistics
    total_lines = fields.Integer(string='Total Lines')
    imported_lines = fields.Integer(string='Imported Lines')
    error_lines = fields.Integer(string='Error Lines')

    # Details
    error_messages = fields.Text(string='Error Messages')
    notes = fields.Text(string='Notes')

    @api.depends('project_id', 'filename', 'import_date')
    def _compute_display_name(self):
        for record in self:
            record.display_name = f"{record.project_id.name} - {record.filename} ({record.import_date.strftime('%Y-%m-%d %H:%M') if record.import_date else 'N/A'})"


class SmetaImportProcessor(models.TransientModel):
    """Helper model for processing smeta imports"""
    _name = 'construction.smeta.processor'
    _description = 'Smeta Import Processor'

    def _get_excel_data(self, file_content, filename):
        """Extract data from Excel file"""
        if not xlrd and not openpyxl:
            raise UserError(_("Please install 'xlrd' and 'openpyxl' Python packages to import Excel files."))

        file_extension = filename.lower().split('.')[-1]

        try:
            if file_extension in ['xls']:
                # Handle old Excel format with xlrd
                if not xlrd:
                    raise UserError(_("Please install 'xlrd' Python package to import .xls files."))

                workbook = xlrd.open_workbook(file_contents=file_content)
                sheet = workbook.sheet_by_index(0)

                data = []
                for row_idx in range(sheet.nrows):
                    row_data = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row_data.append(str(cell_value) if cell_value is not None else '')
                    data.append(row_data)

                return data

            elif file_extension in ['xlsx', 'xlsm']:
                # Handle new Excel format with openpyxl
                if not openpyxl:
                    raise UserError(_("Please install 'openpyxl' Python package to import .xlsx files."))

                workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
                sheet = workbook.active

                data = []
                for row in sheet.iter_rows(values_only=True):
                    row_data = []
                    for cell_value in row:
                        row_data.append(str(cell_value) if cell_value is not None else '')
                    data.append(row_data)

                workbook.close()
                return data

            else:
                raise UserError(_("Unsupported file format. Please upload .xls or .xlsx files."))

        except Exception as e:
            _logger.error(f"Error reading Excel file {filename}: {str(e)}")
            raise UserError(_("Error reading Excel file: %s") % str(e))

    def _clean_numeric_value(self, value):
        """Clean and convert value to float"""
        if not value or value == '':
            return 0.0

        # Convert to string and clean
        str_value = str(value).replace(',', '.').replace(' ', '').replace('\xa0', '')

        # Remove currency symbols and other non-numeric characters except digits, dots, and minus
        import re
        cleaned = re.sub(r'[^\d\.\-]', '', str_value)

        try:
            return float(cleaned) if cleaned else 0.0
        except ValueError:
            return 0.0

    def _find_or_create_category(self, category_name, project_id):
        """Find existing category or create new one"""
        if not category_name or category_name.strip() == '':
            # Get default category
            default_category = self.env['construction.budget.category'].search([
                ('name', 'ilike', 'materials')
            ], limit=1)
            if not default_category:
                default_category = self.env['construction.budget.category'].search([], limit=1)
            return default_category

        # Look for existing category (case insensitive)
        category = self.env['construction.budget.category'].search([
            ('name', 'ilike', category_name.strip())
        ], limit=1)

        if not category:
            # Create new category
            category = self.env['construction.budget.category'].create({
                'name': category_name.strip(),
                'description': f'Auto-created from smeta import for project {project_id}'
            })
            _logger.info(f"Created new budget category: {category_name}")

        return category

    def _find_or_create_uom(self, uom_name):
        """Find existing UoM or create new one"""
        if not uom_name or uom_name.strip() == '':
            # Return default UoM (Units)
            return self.env.ref('uom.product_uom_unit', raise_if_not_found=False) or self.env['uom.uom'].search([], limit=1)

        uom_name = uom_name.strip()

        # Common UoM mappings
        uom_mapping = {
            'шт': 'Units',
            'штук': 'Units',
            'штука': 'Units',
            'pieces': 'Units',
            'pcs': 'Units',
            'м': 'm',
            'метр': 'm',
            'meters': 'm',
            'м2': 'm²',
            'кв.м': 'm²',
            'м²': 'm²',
            'sqm': 'm²',
            'м3': 'm³',
            'куб.м': 'm³',
            'м³': 'm³',
            'cbm': 'm³',
            'кг': 'kg',
            'килограмм': 'kg',
            'kg': 'kg',
            'т': 't',
            'тонна': 't',
            'ton': 't',
            'л': 'l',
            'литр': 'l',
            'liter': 'l',
            'час': 'Hours',
            'ч': 'Hours',
            'hour': 'Hours',
            'hours': 'Hours',
        }

        # Try to find mapped name first
        search_name = uom_mapping.get(uom_name.lower(), uom_name)

        # Search for existing UoM
        uom = self.env['uom.uom'].search([
            '|', ('name', 'ilike', search_name),
            ('name', 'ilike', uom_name)
        ], limit=1)

        if not uom:
            # Create new UoM in 'Units' category
            uom_category = self.env.ref('uom.product_uom_categ_unit', raise_if_not_found=False)
            if not uom_category:
                uom_category = self.env['uom.category'].search([], limit=1)

            uom = self.env['uom.uom'].create({
                'name': uom_name,
                'category_id': uom_category.id,
                'factor': 1.0,
                'uom_type': 'reference',
            })
            _logger.info(f"Created new UoM: {uom_name}")

        return uom

    def _parse_russian_smeta_structure(self, excel_data):
        """Parse Russian smeta structure with sections, main tasks, and sub-tasks"""
        parsed_data = []
        current_section = ''
        current_main_task = None

        for row_idx, row_data in enumerate(excel_data):
            if len(row_data) < 3:
                continue

            # Clean row data
            row_clean = [str(cell).strip() for cell in row_data]

            # Skip header rows and empty rows
            if not any(row_clean) or row_idx < 3:
                continue

            # Check for section header (РАЗДЕЛ:)
            if 'РАЗДЕЛ:' in row_clean[0]:
                current_section = row_clean[0].replace('РАЗДЕЛ:', '').strip()
                continue

            # Check for main task (single number like "1", "2", "3")
            if (row_clean[0] and
                row_clean[0].replace('.', '').isdigit() and
                '.' not in row_clean[0] and
                len(row_clean[2]) > 20):  # Has substantial description

                current_main_task = {
                    'number': row_clean[0],
                    'justification': row_clean[1] if len(row_clean) > 1 else '',
                    'name': row_clean[2] if len(row_clean) > 2 else '',
                    'unit': row_clean[3] if len(row_clean) > 3 else '',
                    'quantity_per_unit': row_clean[4] if len(row_clean) > 4 else '',
                    'total_quantity': row_clean[5] if len(row_clean) > 5 else '',
                    'section': current_section,
                    'type': 'main_task',
                    'parent_task': None,
                    'sub_tasks': []
                }
                parsed_data.append(current_main_task)
                continue

            # Check for sub-task (decimal number like "1.1", "1.2", "2.1")
            if (row_clean[0] and
                '.' in row_clean[0] and
                row_clean[0].replace('.', '').isdigit() and
                len(row_clean[2]) > 5):  # Has description

                sub_task = {
                    'number': row_clean[0],
                    'justification': row_clean[1] if len(row_clean) > 1 else '',
                    'name': row_clean[2] if len(row_clean) > 2 else '',
                    'unit': row_clean[3] if len(row_clean) > 3 else '',
                    'quantity_per_unit': row_clean[4] if len(row_clean) > 4 else '',
                    'total_quantity': row_clean[5] if len(row_clean) > 5 else '',
                    'section': current_section,
                    'type': 'sub_task',
                    'parent_task': current_main_task['number'] if current_main_task else None,
                }

                if current_main_task:
                    current_main_task['sub_tasks'].append(sub_task)
                parsed_data.append(sub_task)

        return parsed_data

    def process_smeta_data(self, excel_data, column_mapping, project_id, budget_id=None):
        """Process the Russian smeta Excel data and create hierarchical budget lines"""
        if not excel_data or len(excel_data) < 4:
            raise UserError(_("Excel file must contain header rows and data."))

        project = self.env['project.project'].browse(project_id)
        if not project.exists():
            raise UserError(_("Selected project does not exist."))

        # Get or create budget
        if budget_id:
            budget = self.env['construction.project.budget'].browse(budget_id)
            if not budget.exists():
                raise UserError(_("Selected budget does not exist."))
        else:
            # Create new budget
            budget = self.env['construction.project.budget'].create({
                'name': f'Imported Smeta - {datetime.now().strftime("%Y-%m-%d %H:%M")}',
                'project_id': project_id,
                'start_date': fields.Date.today(),
            })

        # Create import log
        import_log = self.env['construction.smeta.import.log'].create({
            'project_id': project_id,
            'budget_id': budget.id,
            'filename': 'Russian Smeta Import',
            'state': 'processing',
            'total_lines': len(excel_data),
        })

        try:
            # Parse Russian smeta structure
            parsed_tasks = self._parse_russian_smeta_structure(excel_data)

            errors = []
            imported_count = 0
            budget_lines_to_create = []

            for task in parsed_tasks:
                try:
                    # Only create budget lines for sub-tasks (they have actual quantities)
                    if task['type'] == 'sub_task' and task['total_quantity']:
                        quantity = self._clean_numeric_value(task['total_quantity'])

                        if quantity > 0:  # Only import items with actual quantities
                            # Find or create category based on section
                            category = self._find_or_create_category(task['section'], project.name)
                            uom = self._find_or_create_uom(task['unit'])

                            # Create descriptive name including parent task context
                            if task['parent_task']:
                                item_name = f"{task['parent_task']}. {task['number']} - {task['name']}"
                            else:
                                item_name = f"{task['number']} - {task['name']}"

                            # Limit name length for database
                            if len(item_name) > 200:
                                item_name = item_name[:197] + '...'

                            budget_line_data = {
                                'budget_id': budget.id,
                                'project_id': project_id,
                                'category_id': category.id,
                                'name': item_name,
                                'quantity': quantity,
                                'uom_id': uom.id,
                                'unit_price': 0.0,  # No unit price in smeta format
                                'budget_amount': 0.0,  # Will be filled when actual costs are recorded
                                'sequence': int(task['number'].replace('.', '')) if task['number'].replace('.', '').isdigit() else 99,
                            }

                            budget_lines_to_create.append(budget_line_data)
                            imported_count += 1

                except Exception as e:
                    error_msg = f"Task {task.get('number', 'unknown')}: {str(e)}"
                    errors.append(error_msg)
                    _logger.warning(error_msg)

            # Create all budget lines
            if budget_lines_to_create:
                self.env['construction.project.budget.line'].create(budget_lines_to_create)

            # Update import log
            import_log.write({
                'state': 'success' if not errors else 'failed',
                'imported_lines': imported_count,
                'error_lines': len(errors),
                'error_messages': '\n'.join(errors) if errors else '',
                'notes': f'Successfully imported {imported_count} budget lines (sub-tasks) from Russian smeta into budget: {budget.name}. Parsed {len(parsed_tasks)} total tasks.'
            })

            return {
                'budget_id': budget.id,
                'imported_count': imported_count,
                'error_count': len(errors),
                'errors': errors,
                'import_log_id': import_log.id,
                'parsed_tasks': len(parsed_tasks),
            }

        except Exception as e:
            import_log.write({
                'state': 'failed',
                'error_messages': str(e),
            })
            raise